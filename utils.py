from config import INPUT_DIR, OUTPUT_DIR, TEXT_DATA_FOR_ROW_TWO, MONTH_YEAR_CELL, summary, loan_book_movement, collections_and_overdues, prepayments_and_reschedulement
import os
from openpyxl.styles import Font, PatternFill
from colors import *
from openpyxl.utils import get_column_letter

def listdir_nohidden(path):
    for f in os.listdir(path):
        if not f.startswith('.'):
            yield f

def get_sheets_in_dir(dir):
    return listdir_nohidden(dir)

def get_visible_sheet_list(wb):
    sheet_list = []
    for sheet in wb.sheetnames:
        if wb[sheet].sheet_state == 'visible':
            sheet_list.append(sheet)
    return sheet_list

def get_sheet_row_count(ws):
    if ws.title in [loan_book_movement, collections_and_overdues]:
        col = "B" if ws.title == collections_and_overdues else "D"
        for cell in ws[col]:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.lower().startswith("total"):
                if cell.row + 10 < ws.max_row:
                    return cell.row
    return ws.max_row

def get_sheet_column(ws):
    if ws.title in [loan_book_movement, collections_and_overdues]:
        row = 4 if ws.title == collections_and_overdues else 5
        for cell in ws[row]:
            if cell.value is None:
                if cell.column + 10 < ws.max_column:
                    if ws[row][cell.column + 1].value is None and ws[row][cell.column + 2].value is None:
                        return cell.column
    return ws.max_column

def update_font(ws, size, max_row, max_col):
    try:
        for rows in ws.iter_rows(min_row=1, min_col=1, max_row=max_row ,max_col=max_col):
            for index, cell in enumerate(rows):
                name = cell.font.name
                charset = cell.font.charset
                family = cell.font.family
                b = cell.font.b
                i = cell.font.i
                strike = cell.font.strike
                outline = cell.font.outline
                shadow = cell.font.shadow
                condense = cell.font.condense
                color = cell.font.color
                cell.font = Font(name=name, charset=charset, family=family, b=b, i=i, strike=strike, outline=outline, shadow=shadow, condense=condense, size=size, color=color)

                patternType = cell.fill.patternType
                cell.fill = PatternFill(patternType=patternType, fgColor="FFFFFFFF")

                if index in [0, 1, 2]:
                    cell.font = Font(name=name, charset=charset, family=family, b=True, i=i, strike=strike, outline=outline, shadow=shadow, condense=condense, size=size, color=color)
    except Exception as e:
        print_bold_red(f"Unable to update font: {e}")

def insert_row_a2(ws, a2, bankname, sheet, max_col):
    try:
        col = get_column_letter(max_col)
        ws.move_range(f"A2:{col}{ws.max_row}", rows=1, translate=True)
        ws["A2"].value = f"{TEXT_DATA_FOR_ROW_TWO} {a2} {bankname}"
        if sheet.startswith("Loan") or sheet == prepayments_and_reschedulement:
            ws.delete_rows(4) 
        if sheet == collections_and_overdues:
            ws.delete_rows(3)
    except Exception as e:
        print_bold_red(f"Unable to insert row: {e}")

def move_files_to_output_folder(files):
    for file in files:
        source = os.path.join(INPUT_DIR, file)
        des = os.path.join(OUTPUT_DIR, file)
        os.rename(source, des)
        
def get_year_and_month_for_a2(wb):
    sheet = wb.sheetnames[0]     
    if summary in wb.sheetnames:
       sheet = summary
        
    a2_data = wb[sheet][MONTH_YEAR_CELL].value
    if a2_data:
        data = "{:%B %Y}".format(a2_data)
        print_bold_header(f"Year and Month: {data}")
        return data    
    print_bold_red(f"Year and Month Missing in cell {MONTH_YEAR_CELL} of sheet {sheet}")
    return a2_data

def get_bankname(file):
    l = file.split(" ")
    bankname = f"{l[2]} {l[4]} {l[5]}"
    print_bold_header(f"Bank name to be appended in inserted row {bankname}")
    return bankname