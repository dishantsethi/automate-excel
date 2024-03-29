from config import PDF_DIR, OUTPUT_DIR, summary, loan_book_movement, prepayments_and_reschedulement, collections_and_overdues, PAGE_SCALE
from utils import get_sheets_in_dir, get_visible_sheet_list, get_sheet_row_count, get_sheet_column
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
import os
from openpyxl.styles import Border
from colors import print_bold_header, print_bold_blue, print_bold_warning, print_bold_green, print_bold_red
import win32com.client
from pywintypes import com_error
import time
import shutil

def convert_to_pdf():

    generate_temp_excel_for_pdf()
    update_temp_excel_and_convert_to_pdf()
    

def generate_temp_excel_for_pdf():
    for file in get_sheets_in_dir(OUTPUT_DIR):
        source = os.path.join(OUTPUT_DIR, file)
        des = os.path.join(PDF_DIR, f"tmp{file}")
        shutil.copy(source, des)

def update_temp_excel_and_convert_to_pdf():
    for file in get_sheets_in_dir(PDF_DIR):
        print_bold_blue("------------------------------------------------------------")
        tic = time.time()
        if file.endswith(".pdf"):
            continue
        filename = os.path.join(PDF_DIR, file)
        wb = load_workbook(filename=filename)
        sheet_list = get_visible_sheet_list(wb)
        for sheet in sheet_list:
            ws = wb[sheet]
            ws.oddFooter.center.text = 'Page &P of &N' 
            ws.evenFooter.center.text = 'Page &P of &N'
            if sheet == summary or sheet.startswith(summary):
                ws.page_margins = PageMargins(left=0.50, right=0.50, top=0.50, bottom=1.50, header=0.3, footer=0.3)
                ws.page_setup = PrintPageSetup(orientation=ws.ORIENTATION_PORTRAIT, paperSize=ws.PAPERSIZE_A4, scale=PAGE_SCALE[summary])
            elif sheet == loan_book_movement or sheet.startswith("Loan"):
                ws.page_margins = PageMargins(left=0.50, right=0.50, top=0.50, bottom=1.50, header=0.3, footer=0.3)
                ws.page_setup = PrintPageSetup(orientation=ws.ORIENTATION_LANDSCAPE, paperSize=ws.PAPERSIZE_A4, scale=PAGE_SCALE[loan_book_movement])
                ws.print_title_cols = 'A:D'
                ws.print_title_rows = '1:5'
                update_border(ws)
            elif sheet == prepayments_and_reschedulement:
                ws.page_margins = PageMargins(left=0.50, right=0.50, top=0.50, bottom=1.50, header=0.3, footer=0.3)
                ws.page_setup = PrintPageSetup(orientation=ws.ORIENTATION_LANDSCAPE, paperSize=ws.PAPERSIZE_A4, scale=PAGE_SCALE[prepayments_and_reschedulement])
            elif sheet == collections_and_overdues:
                ws.page_margins = PageMargins(left=0.50, right=0.50, top=0.50, bottom=1.50, header=0.3, footer=0.3)
                ws.page_setup = PrintPageSetup(orientation=ws.ORIENTATION_LANDSCAPE, paperSize=ws.PAPERSIZE_A4, scale=PAGE_SCALE[collections_and_overdues])
                ws.print_title_cols = 'A:B'
                ws.print_title_rows = '1:4'
                update_border(ws)
            else:
                ws.page_margins = PageMargins(left=0.50, right=0.50, top=0.50, bottom=1.50, header=0.3, footer=0.3)
                ws.page_setup = PrintPageSetup(orientation=ws.ORIENTATION_LANDSCAPE, paperSize=ws.PAPERSIZE_A4, scale=PAGE_SCALE["default"])
        wb.save(filename)
        print_bold_header(f"Page Setup Done for file {file}")
        print_bold_header(f"Borders removed for file {file}")
        excel_to_pdf(file, sheet_list)
        toc = time.time()
        print_bold_green(f"Time Take: {toc-tic} seconds")
        print_bold_blue("------------------------------------------------------------")


def update_border(ws):
    max_row = get_sheet_row_count(ws)
    max_col = get_sheet_column(ws)
    for rows in ws.iter_rows(min_row=1, min_col=1, max_row=max_row ,max_col=max_col):
        for cell in rows:
            cell.border = Border(left=None, right=None, bottom=None, top=None, outline=None)
        
def excel_to_pdf(excel_file, sheet_list):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.interactive = False
    excel.visible = False
    pdf_file = excel_file[3:len(excel_file)-4] + "pdf"
    pdf_file_path = os.path.join(PDF_DIR, pdf_file)
    excel_file_path = os.path.join(PDF_DIR, excel_file)
    try:
        wb = excel.Workbooks.Open(excel_file_path)
        sheet_count = sheet_list.index(collections_and_overdues) + 1
        ws_index_list = [x for x in range(1, sheet_count+1)]
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, pdf_file_path)
    except com_error as e:
        print_bold_red(f'Failed. {e}')
    except Exception as e:
        print_bold_red(f"Error: {e}")
    else:
        print_bold_warning(f"Converted {excel_file} to {pdf_file}")
    finally:
        wb.Close()
        excel.Quit()
    
